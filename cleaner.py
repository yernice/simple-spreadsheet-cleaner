import openpyxl as xl
import utils
import logging


logging.basicConfig(
    filename='spreadsheet_cleaner.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)


wb_input = xl.load_workbook('messy_spreadsheet.xlsx')
raw_data = wb_input['Sheet1']
wb_output = xl.Workbook()
clean_data = wb_output.active
clean_data.title = "CleanSheet"

current_row = 1
for row in range(1, raw_data.max_row + 1):
    if row == 1:
        utils.copy_first_row(raw_data, clean_data)
        current_row += 1
        continue

    if utils.skip(raw_data, row):
        logging.warning(f"Row {row} was skipped due to uncomplete data.")
        continue

    if utils.duplicate(raw_data, clean_data, row, current_row):
        logging.warning(f"Row {row} was skipped due to duplicate data.")
        continue

    name = raw_data.cell(row=row, column=1).value
    name = utils.clean_value(name)
    mail = raw_data.cell(row=row, column=2).value
    mail = utils.clean_value(mail)
    age = raw_data.cell(row=row, column=3).value

    clean_data.cell(row=current_row, column=1).value = name
    clean_data.cell(row=current_row, column=2).value = mail
    clean_data.cell(row=current_row, column=3).value = age
    current_row += 1

    logging.info(f"Row {row} was succesfully cleaned.")


wb_output.save("cleaned_spreadsheet.xlsx")
